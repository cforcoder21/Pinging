import sys, os, subprocess, json

# ── venv bootstrap (cross-platform) ──────────────────────────────────────────
_base = os.path.dirname(os.path.abspath(__file__))
if sys.platform == "win32":
    _venv = os.path.join(_base, ".venv", "Lib", "site-packages")
else:
    _ver  = f"python{sys.version_info.major}.{sys.version_info.minor}"
    _venv = os.path.join(_base, ".venv", "lib", _ver, "site-packages")
if os.path.exists(_venv) and _venv not in sys.path:
    sys.path.insert(0, _venv)

import csv, math, threading, socket
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

# ── cross-platform ping ───────────────────────────────────────────────────────

def _subprocess_ping(ip):
    try:
        if sys.platform == "win32":
            cmd = ["ping", "-n", "1", "-w", "800", ip]
        else:
            cmd = ["ping", "-c", "1", "-W", "800", "-t", "1", ip]
        r = subprocess.run(cmd, stdout=subprocess.PIPE,
                           stderr=subprocess.DEVNULL, timeout=3)
        if r.returncode == 0:
            out = r.stdout.decode(errors="ignore")
            for token in out.split():
                t = token.lower().replace("time=", "").replace("time<", "") \
                         .replace("ms", "").rstrip("<")
                try:
                    return float(t)
                except ValueError:
                    pass
            return 1.0
    except Exception:
        pass
    return None

try:
    from ping3 import ping as _ping3
    def do_ping(ip, timeout=0.8):
        try:
            r = _ping3(ip, timeout=timeout)
            return round(r * 1000, 1) if r else None
        except PermissionError:
            return _subprocess_ping(ip)
        except Exception:
            return _subprocess_ping(ip)
except ImportError:
    def do_ping(ip, timeout=0.8):
        return _subprocess_ping(ip)

# ── config ────────────────────────────────────────────────────────────────────

def resource(name):
    return os.path.join(getattr(sys, "_MEIPASS", os.path.abspath(".")), name)

CSV_FILE     = resource("coop_locations_large.csv")
LAYOUT_FILE  = os.path.join(_base, "layout.json")
MAX_THREADS  = 60
PANEL_W      = 295
STRIP_W      = 26

# ── colour palette ────────────────────────────────────────────────────────────

BG_MAP    = "#f0f4fa"
BG_PANEL  = "#f8f9fc"
BG_WIDGET = "#ffffff"
FG_TEXT   = "#1a1f36"
FG_SUB    = "#6b7280"
ACCENT    = "#2563eb"
SEP_COL   = "#e5e7eb"

C_ONLINE_CORE  = "#16a34a"
C_ONLINE_MID   = "#4ade80"
C_ONLINE_GLOW  = "#dcfce7"
C_OFFLINE_CORE = "#dc2626"
C_OFFLINE_GLOW = "#fee2e2"
C_SOURCE_CORE  = "#2563eb"
C_SOURCE_MID   = "#60a5fa"
C_SOURCE_GLOW  = "#dbeafe"
C_LINE_ON      = "#bbf7d0"
C_LINE_OFF     = "#fecaca"

DEPT_COLORS = [
    "#3b82f6","#8b5cf6","#f97316","#06b6d4",
    "#eab308","#ef4444","#22c55e","#6b7280",
    "#ec4899","#0ea5e9",
]

LOCK_ON_BG  = "#fef3c7"
LOCK_ON_FG  = "#92400e"
LOCK_OFF_BG = "#f3f4f6"
LOCK_OFF_FG = "#374151"

# ── colour helpers ────────────────────────────────────────────────────────────

def hex_lighter(h, amt=0.55):
    h = h.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = min(255, int(r + (255 - r) * amt))
    g = min(255, int(g + (255 - g) * amt))
    b = min(255, int(b + (255 - b) * amt))
    return f"#{r:02x}{g:02x}{b:02x}"

def hex_darker(h, amt=0.25):
    h = h.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = max(0, int(r * (1 - amt)))
    g = max(0, int(g * (1 - amt)))
    b = max(0, int(b * (1 - amt)))
    return f"#{r:02x}{g:02x}{b:02x}"

# ── data helpers ──────────────────────────────────────────────────────────────

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80)); ip = s.getsockname()[0]; s.close()
        return ip
    except Exception:
        return "127.0.0.1"

def load_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls") and HAS_XLSX:
        return _from_xlsx(path)
    return _from_csv(path)

def _from_csv(path):
    rows = []
    if not os.path.exists(path): return rows
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            ip = r.get("IP", "").strip()
            if ip:
                rows.append({"IP": ip,
                             "Name": r.get("Name", "").strip(),
                             "Department": r.get("Department", "Unknown").strip(),
                             "Location": r.get("Location", "").strip(),
                             "Parent IP": r.get("Parent IP", "").strip(),
                             "Zone": r.get("Zone", "").strip()})
    return rows

def _from_xlsx(path):
    rows = []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    hdrs = [str(c.value or "").strip() for c in next(ws.iter_rows(1, 1))]
    col  = {h: i for i, h in enumerate(hdrs)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ip = str(row[col.get("IP", 0)] or "").strip()
        if ip:
            rows.append({"IP": ip,
                         "Name": str(row[col.get("Name", 1)] or "").strip(),
                         "Department": str(row[col.get("Department", 2)] or "Unknown").strip(),
                         "Location": str(row[col.get("Location", 3)] or "").strip() if "Location" in col else "",
                         "Parent IP": str(row[col.get("Parent IP", -1)] or "").strip() if "Parent IP" in col else "",
                         "Zone": str(row[col.get("Zone", -1)] or "").strip() if "Zone" in col else ""})
    return rows

def _tree_layout(source_ip, devices):
    """Left-to-right tree: source at col 0, children at col+1, spread vertically."""
    H_STEP = 195
    V_STEP = 60

    all_ips  = set(d["IP"] for d in devices) | {source_ip}
    children = {ip: [] for ip in all_ips}

    for d in devices:
        ip  = d["IP"]
        pid = (d.get("Parent IP") or "").strip()
        if pid and pid in all_ips:
            children[pid].append(ip)
        else:
            children[source_ip].append(ip)

    def leaf_count(ip):
        ch = children.get(ip, [])
        return max(1, sum(leaf_count(c) for c in ch))

    positions = {}

    def layout(ip, col, row_offset, row_span):
        positions[ip] = (col * H_STEP, (row_offset + row_span / 2) * V_STEP)
        ch = children.get(ip, [])
        r  = row_offset
        for c in ch:
            span = leaf_count(c)
            layout(c, col + 1, r, span)
            r += span

    total = leaf_count(source_ip)
    layout(source_ip, 0, 0, total)
    sx, sy = positions.get(source_ip, (0, 0))
    return {ip: (x - sx, y - sy) for ip, (x, y) in positions.items()}

def sunflower(n, cx, cy, spacing=28):
    golden = (1 + math.sqrt(5)) / 2
    pts = []
    for i in range(n):
        r     = spacing * math.sqrt(i + 0.5)
        theta = 2 * math.pi * i / golden ** 2
        pts.append((cx + r * math.cos(theta), cy + r * math.sin(theta)))
    return pts

# ── tooltip ───────────────────────────────────────────────────────────────────

class Tooltip:
    def __init__(self, canvas):
        self._cv  = canvas
        self._win = None

    def show(self, text, rx, ry):
        self.hide()
        self._win = tw = tk.Toplevel(self._cv)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{rx + 16}+{ry - 12}")
        outer = tk.Frame(tw, bg="#1e293b", padx=1, pady=1)
        outer.pack()
        tk.Label(outer, text=text, justify="left",
                 bg="#1e293b", fg="#f8fafc",
                 padx=10, pady=8, font=("Arial", 10),
                 relief="flat").pack()

    def hide(self):
        if self._win:
            try: self._win.destroy()
            except Exception: pass
            self._win = None

# ── map canvas ────────────────────────────────────────────────────────────────

class MapCanvas(tk.Canvas):
    R_SRC = 22
    R_DEV = 10

    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAP, highlightthickness=0)
        self._zoom     = 1.0
        self._pan      = [0.0, 0.0]
        self._drag_xy  = None
        self._drag_ip  = None
        self._drag_off = (0, 0)
        self._locked   = False
        self._nodes    = {}
        self._hovered  = None
        self._tip      = Tooltip(self)

        self.bind("<ButtonPress-1>",   self._b1_down)
        self.bind("<B1-Motion>",       self._b1_drag)
        self.bind("<ButtonRelease-1>", self._b1_up)
        self.bind("<MouseWheel>",      self._wheel)
        self.bind("<Button-4>",        self._wheel)
        self.bind("<Button-5>",        self._wheel)
        self.bind("<Configure>",       lambda _: self._redraw())
        self.bind("<Motion>",          self._mouse_move)
        self.bind("<Leave>",           lambda _: self._tip.hide())

    def set_locked(self, locked):
        self._locked = locked
        self.configure(cursor="hand2" if locked else "fleur")

    def _tc(self, lx, ly):
        w = self.winfo_width()  or 900
        h = self.winfo_height() or 600
        return (w/2 + (lx + self._pan[0]) * self._zoom,
                h/2 + (ly + self._pan[1]) * self._zoom)

    def _tl(self, cx, cy):
        w = self.winfo_width()  or 900
        h = self.winfo_height() or 600
        return ((cx - w/2) / self._zoom - self._pan[0],
                (cy - h/2) / self._zoom - self._pan[1])

    def _hit(self, ex, ey):
        hits = self.find_overlapping(ex - 8, ey - 8, ex + 8, ey + 8)
        for ip, d in self._nodes.items():
            if d.get("oval_id") in hits:
                return ip
        return None

    def _b1_down(self, e):
        self._tip.hide()
        ip = self._hit(e.x, e.y)
        if self._locked and ip:
            self._drag_ip  = ip
            self._drag_off = (e.x, e.y)
            self._drag_xy  = None
        elif not self._locked:
            self._drag_ip = None
            self._drag_xy = (e.x, e.y)

    def _b1_drag(self, e):
        if self._locked and self._drag_ip:
            dx = (e.x - self._drag_off[0]) / self._zoom
            dy = (e.y - self._drag_off[1]) / self._zoom
            self._nodes[self._drag_ip]["lx"] += dx
            self._nodes[self._drag_ip]["ly"] += dy
            self._drag_off = (e.x, e.y)
            self._redraw()
        elif not self._locked and self._drag_xy:
            dx = (e.x - self._drag_xy[0]) / self._zoom
            dy = (e.y - self._drag_xy[1]) / self._zoom
            self._pan[0] += dx
            self._pan[1] += dy
            self._drag_xy = (e.x, e.y)
            self._redraw()

    def _b1_up(self, e):
        self._drag_ip = None
        self._drag_xy = None

    def _wheel(self, e):
        delta = getattr(e, "delta", 0) or (120 if e.num == 4 else -120)
        f = 1.15 if delta > 0 else 1 / 1.15
        lx, ly = self._tl(e.x, e.y)
        self._zoom *= f
        w = self.winfo_width()  or 900
        h = self.winfo_height() or 600
        self._pan[0] = (e.x - w/2) / self._zoom - lx
        self._pan[1] = (e.y - h/2) / self._zoom - ly
        self._redraw()

    def fit_view(self):
        if not self._nodes: return
        xs = [d["lx"] for d in self._nodes.values()]
        ys = [d["ly"] for d in self._nodes.values()]
        self._pan = [-(min(xs)+max(xs))/2, -(min(ys)+max(ys))/2]
        w = self.winfo_width()  or 900
        h = self.winfo_height() or 600
        self._zoom = min(w / max(max(xs)-min(xs)+240, 1),
                         h / max(max(ys)-min(ys)+240, 1), 1.8)
        self._redraw()

    def _mouse_move(self, e):
        ip = self._hit(e.x, e.y)
        if ip != self._hovered:
            self._hovered = ip
            if ip:
                d   = self._nodes[ip]
                lat = d.get("latency")
                tip = (f"IP:        {ip}\n"
                       f"Name:      {d['name']}\n"
                       f"Dept:      {d['dept']}\n"
                       f"Location:  {d.get('location') or '—'}\n"
                       f"Status:    {'Online  ' + str(lat) + ' ms' if lat else 'Offline'}")
                self._tip.show(tip, self.winfo_rootx()+e.x, self.winfo_rooty()+e.y)
            else:
                self._tip.hide()

    def set_nodes(self, node_list, preserved_positions=None):
        old = {ip: (d["lx"], d["ly"]) for ip, d in self._nodes.items()}
        if preserved_positions:
            old.update(preserved_positions)
        self._nodes = {}
        for d in node_list:
            ip = d["ip"]
            lx, ly = old.get(ip, (d["lx"], d["ly"]))
            self._nodes[ip] = {
                "lx": lx, "ly": ly,
                "fill": d["fill"],
                "name": d["name"],
                "dept": d["dept"],
                "location": d.get("location", ""),
                "is_source": d.get("is_source", False),
                "parent_ip": d.get("parent_ip"),
                "zone": d.get("zone", ""),
                "latency": None,
                "oval_id": None,
            }
        self._redraw()

    def update_status(self, ip, latency):
        if ip in self._nodes:
            self._nodes[ip]["latency"] = latency

    # ── drawing ───────────────────────────────────────────────────────────────

    def _box_size(self, d):
        """Return (half_w, half_h) of a node's rectangle at current zoom."""
        z    = self._zoom
        name = "YOU" if d["is_source"] else d["name"]
        fs   = max(7, int((10 if d["is_source"] else 9) * z))
        bw   = max(44 * z, len(name) * fs * 0.63 + 20 * z)
        bh   = max(13 * z, fs * 0.5 + 9 * z)
        return bw, bh

    def _redraw(self):
        self.delete("all")
        if not self._nodes: return
        w = self.winfo_width()  or 900
        h = self.winfo_height() or 600

        # pre-compute box sizes so line routing can use them
        for d in self._nodes.values():
            d["_bw"], d["_bh"] = self._box_size(d)

        # ── dot grid ─────────────────────────────────────────────────────────
        step = max(30, int(45 * self._zoom))
        ox   = int((self._pan[0] * self._zoom + w/2) % step)
        oy   = int((self._pan[1] * self._zoom + h/2) % step)
        for gx in range(ox - step, w + step, step):
            for gy in range(oy - step, h + step, step):
                self.create_oval(gx-1, gy-1, gx+1, gy+1, fill="#dde3ec", outline="")

        # source position
        src  = next((d for d in self._nodes.values() if d["is_source"]), None)
        sx, sy = self._tc(src["lx"], src["ly"]) if src else self._tc(0, 0)

        # ── zone boxes (only for nodes that have an explicit Zone value) ────────
        _ZONE_BORDERS = ["#1a1a2e", "#0f766e", "#1e40af", "#6b21a8",
                         "#9a3412", "#065f46", "#7c2d12"]
        zone_nodes: dict[str, list] = {}
        for ip, d in self._nodes.items():
            z_name = d.get("zone", "")
            if z_name:
                zone_nodes.setdefault(z_name, []).append(d)

        for zi, (zone_name, znodes) in enumerate(zone_nodes.items()):
            coords = [self._tc(n["lx"], n["ly"]) for n in znodes]
            cxs    = [c[0] for c in coords]
            cys    = [c[1] for c in coords]
            ebw    = max(n["_bw"] for n in znodes)
            ebh    = max(n["_bh"] for n in znodes)
            pad    = max(22, 30 * self._zoom)
            x0     = min(cxs) - ebw - pad
            y0     = min(cys) - ebh - pad
            x1     = max(cxs) + ebw + pad
            y1     = max(cys) + ebh + pad
            bcol   = _ZONE_BORDERS[zi % len(_ZONE_BORDERS)]
            bw_px  = max(2, self._zoom * 2.2)

            self.create_rectangle(x0, y0, x1, y1,
                                  fill="#fafcff", outline=bcol, width=bw_px)
            fs_z = max(7, int(9 * self._zoom))
            self.create_text(x0 + 8, y0 + 5, text=zone_name, fill="#dc2626",
                             font=("Arial", fs_z, "bold"), anchor="nw")

        # ── topology lines — orthogonal routing ──────────────────────────────
        lw = max(1.0, self._zoom * 1.1)
        for ip, d in self._nodes.items():
            if d["is_source"]: continue
            tx, ty = self._tc(d["lx"], d["ly"])
            tbw, tbh = d["_bw"], d["_bh"]
            pid = d.get("parent_ip")
            if pid and pid in self._nodes:
                p            = self._nodes[pid]
                px_, py_     = self._tc(p["lx"], p["ly"])
                pbw, pbh     = p["_bw"], p["_bh"]
                parent_up    = p["is_source"] or (p.get("latency") is not None)
            else:
                px_, py_     = sx, sy
                pbw          = src["_bw"] if src else 30 * self._zoom
                pbh          = src["_bh"] if src else 13 * self._zoom
                parent_up    = True

            col  = "#0d9488" if parent_up else "#dc2626"
            dash = ()        if parent_up else (6, 5)

            # choose exit/entry edges based on relative position
            dx_  = tx - px_
            dy_  = ty - py_
            if abs(dy_) >= abs(dx_):          # more vertical → top/bottom exit
                if ty >= py_:
                    ax, ay = px_, py_ + pbh   # bottom of parent
                    bx, by_ = tx, ty - tbh    # top of child
                else:
                    ax, ay = px_, py_ - pbh
                    bx, by_ = tx, ty + tbh
                my_ = (ay + by_) / 2
                pts = [ax, ay, ax, my_, bx, my_, bx, by_]
            else:                              # more horizontal → left/right exit
                if tx >= px_:
                    ax, ay = px_ + pbw, py_   # right of parent
                    bx, by_ = tx - tbw, ty    # left of child
                else:
                    ax, ay = px_ - pbw, py_
                    bx, by_ = tx + tbw, ty
                mx_ = (ax + bx) / 2
                pts = [ax, ay, mx_, ay, mx_, by_, bx, by_]

            self.create_line(*pts, fill=col, width=lw, dash=dash,
                             joinstyle="round", capstyle="round")

            # arrowhead at the child end
            self._arrowhead(pts[-4], pts[-3], pts[-2], pts[-1], col, lw)

        # ── nodes ─────────────────────────────────────────────────────────────
        for ip, d in self._nodes.items():
            cx, cy = self._tc(d["lx"], d["ly"])
            self._draw_node(cx, cy, ip, d)

        # ── legend ────────────────────────────────────────────────────────────
        self._legend(14, 14)

        # ── lock banner ───────────────────────────────────────────────────────
        if self._locked:
            self.create_rectangle(0, 0, w, 28,
                                  fill="#fef9c3", outline="#fde047", width=1)
            self.create_text(w//2, 14,
                             text="🔒  LOCK MODE ON  —  drag any node to reposition it",
                             fill="#92400e", font=("Arial", 10, "bold"))

    def _arrowhead(self, x1, y1, x2, y2, col, lw):
        """Draw a small arrowhead at (x2,y2) pointing from (x1,y1)."""
        if x1 == x2 and y1 == y2: return
        size = max(5, lw * 3.5)
        angle = math.atan2(y2 - y1, x2 - x1)
        a1 = angle + math.radians(145)
        a2 = angle - math.radians(145)
        pts = [x2, y2,
               x2 + size * math.cos(a1), y2 + size * math.sin(a1),
               x2 + size * math.cos(a2), y2 + size * math.sin(a2)]
        self.create_polygon(*pts, fill=col, outline="")

    def _draw_node(self, cx, cy, ip, d):
        z      = self._zoom
        is_src = d["is_source"]
        lat    = d.get("latency")
        bw, bh = d["_bw"], d["_bh"]
        name   = "YOU" if is_src else d["name"]
        fs     = max(7, int((10 if is_src else 9) * z))

        if is_src:
            fill_bg = "#eff6ff"
            border  = "#2563eb"
            bord_w  = max(1.5, z * 1.5)
            text_c  = "#1e40af"
            dot_c   = "#2563eb"
        elif lat is not None:
            fill_bg = "white"
            border  = "#d1d5db"
            bord_w  = max(1, z)
            text_c  = "#111827"
            dot_c   = "#16a34a"
        else:
            fill_bg = "white"
            border  = "#e5e7eb"
            bord_w  = max(1, z)
            text_c  = "#9ca3af"
            dot_c   = "#dc2626"

        # Subtle shadow
        self.create_rectangle(cx-bw+1.5, cy-bh+1.5, cx+bw+1.5, cy+bh+1.5,
                              fill="#dde3ec", outline="")
        # Main box — clean white rectangle
        oid = self.create_rectangle(cx-bw, cy-bh, cx+bw, cy+bh,
                                    fill=fill_bg, outline=border, width=bord_w)
        d["oval_id"] = oid

        # Tiny status dot (top-right corner only)
        dr = max(2.5, 3 * z)
        self.create_oval(cx+bw-dr*2-1.5*z, cy-bh+1.5*z,
                         cx+bw-1.5*z,       cy-bh+dr*2+1.5*z,
                         fill=dot_c, outline="white", width=max(0.5, z*0.5))

        # Device name centred in box — no accent bar offset
        self.create_text(cx, cy, text=name, fill=text_c,
                         font=("Arial", fs, "bold" if is_src else "normal"),
                         width=int(bw * 2 - 10))

        # Latency badge when online and zoomed in
        if lat is not None and z > 1.2:
            ms  = f"{lat}ms"
            mfs = max(5, int(6 * z))
            mbw = len(ms) * mfs * 0.72 + 6
            mbh = mfs + 4
            self.create_rectangle(cx-mbw/2, cy+bh+1, cx+mbw/2, cy+bh+mbh+2,
                                  fill="#0d9488", outline="")
            self.create_text(cx, cy+bh+mbh/2+2, text=ms, fill="white",
                             font=("Arial", mfs, "bold"))

        # Source IP label below box
        if is_src:
            self.create_text(cx, cy+bh+5*z, text=ip, fill="#1e40af",
                             font=("Arial", max(6, int(7*z))))

        # Drag handle when locked
        if self._locked and not is_src:
            self.create_oval(cx-2, cy-2, cx+2, cy+2,
                             fill="white", outline="#9ca3af", width=1)

    def _legend(self, x, y):
        """Cable colour indication legend — light blue bordered box, top-left."""
        pad = 10; lh = 20; bw = 196
        items = [
            ("#0d9488", (),      "Path online  (teal solid)"),
            ("#dc2626", (6, 4),  "Path broken  (red dashed)"),
            ("#2563eb", (),      "Source — this machine"),
        ]
        bh = pad * 2 + 17 + lh * len(items)

        # White card with light-blue border (matches reference diagram)
        self.create_rectangle(x, y, x+bw, y+bh,
                              fill="white", outline="#93c5fd",
                              width=max(1.5, self._zoom * 0.8))
        # Title band
        self.create_rectangle(x+1, y+1, x+bw-1, y+18,
                              fill="#eff6ff", outline="")
        self.create_text(x+pad, y+3, text="Cable colour indication",
                         fill="#1e40af", font=("Arial", 8, "bold"), anchor="nw")

        for i, (col, dash, lbl) in enumerate(items):
            oy  = y + pad + 17 + lh * i
            lx1 = x + pad
            lx2 = x + pad + 32
            ly  = oy + lh // 2
            self.create_line(lx1, ly, lx2, ly, fill=col, width=2, dash=dash,
                             capstyle="round")
            self.create_text(lx2 + 7, ly - 1, text=lbl, fill=FG_TEXT,
                             font=("Arial", 8), anchor="w")

# ── main application ──────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Network Map Monitor")
        self.geometry("1380x800")
        self.minsize(900, 560)
        self.configure(bg=BG_MAP)

        self._source_ip   = get_local_ip()
        self._systems     = {}
        self._devices     = []
        self._executor    = ThreadPoolExecutor(max_workers=MAX_THREADS)
        self._search      = tk.StringVar()
        self._dept_var    = tk.StringVar(value="All Departments")
        self._locked      = False
        self._scanning    = False
        self._panel_open  = False
        self._hide_job    = None

        self._search.trace_add("write",   lambda *_: self._refresh_table())
        self._dept_var.trace_add("write", lambda *_: self._refresh_table())

        self._build_ui()
        self._hover_loop()
        self._load(CSV_FILE)
        self._schedule_scan()

    # ── UI layout ─────────────────────────────────────────────────────────────

    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".", background=BG_PANEL, foreground=FG_TEXT,
                        fieldbackground=BG_WIDGET, font=("Arial", 11))
        style.configure("Treeview", background=BG_WIDGET, foreground=FG_TEXT,
                        fieldbackground=BG_WIDGET, rowheight=22)
        style.configure("Treeview.Heading", background="#e5e7eb",
                        foreground=FG_SUB, font=("Arial", 10, "bold"))
        style.map("Treeview", background=[("selected", "#dbeafe")])
        style.configure("TEntry",    fieldbackground=BG_WIDGET, foreground=FG_TEXT)
        style.configure("TCombobox", fieldbackground=BG_WIDGET, foreground=FG_TEXT)
        style.configure("TButton",   background="#e5e7eb", foreground=FG_TEXT, padding=(6,5))
        style.map("TButton", background=[("active", "#d1d5db")])

        # ── strip (always visible, hover target) ──────────────────────────────
        self._strip = tk.Frame(self, bg=ACCENT, width=STRIP_W)
        self._strip.pack(side="left", fill="y")
        self._strip.pack_propagate(False)

        self._arrow = tk.Label(self._strip, text="›", bg=ACCENT, fg="white",
                               font=("Arial", 15, "bold"), cursor="hand2")
        self._arrow.pack(expand=True)

        for w in (self._strip, self._arrow):
            w.bind("<Enter>", lambda e: self._open_panel())

        # ── sliding panel (hidden at width=0 initially) ────────────────────
        self._panel = tk.Frame(self, bg=BG_PANEL, width=0)
        self._panel.pack(side="left", fill="y")
        self._panel.pack_propagate(False)

        # ── map ────────────────────────────────────────────────────────────
        self._map = MapCanvas(self)
        self._map.pack(side="left", fill="both", expand=True)

        # populate the panel content
        self._build_panel()

    def _build_panel(self):
        p = self._panel

        def lbl(text, color=FG_TEXT, size=11, bold=False):
            return tk.Label(p, text=text, bg=BG_PANEL, fg=color,
                            font=("Arial", size, "bold" if bold else "normal"),
                            anchor="w", padx=12)

        lbl("Network Map Monitor", ACCENT, 14, bold=True).pack(fill="x", pady=(12,2))
        self._lbl_src = lbl(f"Source:  {self._source_ip}", FG_SUB, 10)
        self._lbl_src.pack(fill="x")
        ttk.Separator(p).pack(fill="x", pady=8, padx=8)

        self._lbl_stats = lbl("Loading…", FG_TEXT, 11, bold=True)
        self._lbl_stats.pack(fill="x")
        self._lbl_time = lbl("", FG_SUB, 9)
        self._lbl_time.pack(fill="x", pady=(0,6))

        tk.Label(p, text="Search", bg=BG_PANEL, fg=FG_SUB,
                 font=("Arial",9), anchor="w", padx=12).pack(fill="x")
        ttk.Entry(p, textvariable=self._search).pack(fill="x", padx=10, pady=(0,6))

        tk.Label(p, text="Department", bg=BG_PANEL, fg=FG_SUB,
                 font=("Arial",9), anchor="w", padx=12).pack(fill="x")
        self._dept_cb = ttk.Combobox(p, textvariable=self._dept_var, state="readonly")
        self._dept_cb.pack(fill="x", padx=10, pady=(0,8))

        for txt, cmd in [("⟳  Scan Now",         self._scan),
                          ("📂  Load CSV / Excel", self._load_dialog),
                          ("⊡  Fit Map",           lambda: self._map.fit_view()),
                          ("📤  Export CSV",        self._export)]:
            ttk.Button(p, text=txt, command=cmd).pack(fill="x", padx=10, pady=2)

        # Save Layout button
        self._save_btn = tk.Button(
            p, text="💾  Save Layout",
            bg="#e5e7eb", fg=FG_TEXT,
            font=("Arial", 11), relief="flat", bd=0,
            padx=10, pady=5, cursor="hand2",
            command=self._save_layout)
        self._save_btn.pack(fill="x", padx=10, pady=2)

        self._lock_btn = tk.Button(
            p, text="🔓  Lock Map  (OFF)",
            bg=LOCK_OFF_BG, fg=LOCK_OFF_FG,
            font=("Arial", 11, "bold"),
            relief="flat", bd=0, padx=10, pady=6,
            cursor="hand2", command=self._toggle_lock)
        self._lock_btn.pack(fill="x", padx=10, pady=(6,2))

        ttk.Separator(p).pack(fill="x", pady=8, padx=8)
        tk.Label(p, text="Double-click row → jump to node",
                 bg=BG_PANEL, fg=FG_SUB, font=("Arial",9),
                 anchor="w", padx=12).pack(fill="x", pady=(0,4))

        cols = ("IP","Name","Latency","Status")
        self._tree = ttk.Treeview(p, columns=cols, show="headings", selectmode="browse")
        for c in cols:
            self._tree.heading(c, text=c)
            self._tree.column(c, width=60, anchor="center")
        self._tree.column("IP",   width=110, anchor="w")
        self._tree.column("Name", width=100, anchor="w")
        self._tree.tag_configure("online",  background="#dcfce7", foreground="#166534")
        self._tree.tag_configure("offline", background="#fee2e2", foreground="#991b1b")
        sb = ttk.Scrollbar(p, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        self._tree.pack(side="left", fill="both", expand=True, padx=(10,0), pady=(0,10))
        sb.pack(side="left", fill="y", pady=(0,10), padx=(0,6))
        self._tree.bind("<Double-1>", self._jump_to)

    # ── hover panel logic ─────────────────────────────────────────────────────

    def _open_panel(self):
        if self._hide_job:
            self.after_cancel(self._hide_job)
            self._hide_job = None
        if not self._panel_open:
            self._panel.config(width=PANEL_W)
            self._arrow.config(text="‹")
            self._panel_open = True

    def _close_panel(self):
        self._panel.config(width=0)
        self._arrow.config(text="›")
        self._panel_open = False
        self._hide_job   = None

    def _in_panel_zone(self, widget):
        """Return True if widget is the strip, panel, or any descendant of them."""
        w = widget
        while w is not None:
            if w is self._strip or w is self._panel:
                return True
            try:
                w = w.master
            except AttributeError:
                break
        return False

    def _hover_loop(self):
        """Poll every 120 ms — close panel when the cursor leaves the strip+panel zone."""
        if self._panel_open:
            try:
                px  = self.winfo_pointerx()
                py  = self.winfo_pointery()
                hit = self.winfo_containing(px, py)
                if self._in_panel_zone(hit):
                    # cursor is inside — cancel any pending close
                    if self._hide_job:
                        self.after_cancel(self._hide_job)
                        self._hide_job = None
                else:
                    # cursor is outside — schedule close if not already pending
                    if not self._hide_job:
                        self._hide_job = self.after(350, self._close_panel)
            except Exception:
                pass
        self.after(120, self._hover_loop)

    # ── lock ──────────────────────────────────────────────────────────────────

    def _toggle_lock(self):
        self._locked = not self._locked
        self._map.set_locked(self._locked)
        if self._locked:
            self._lock_btn.config(text="🔒  Lock Map  (ON)",
                                  bg=LOCK_ON_BG, fg=LOCK_ON_FG)
        else:
            self._lock_btn.config(text="🔓  Lock Map  (OFF)",
                                  bg=LOCK_OFF_BG, fg=LOCK_OFF_FG)
        self._map._redraw()

    # ── save / load layout ────────────────────────────────────────────────────

    def _save_layout(self):
        positions = {ip: [d["lx"], d["ly"]] for ip, d in self._map._nodes.items()}
        try:
            with open(LAYOUT_FILE, "w") as f:
                json.dump({"version": 1, "positions": positions}, f, indent=2)
            self._save_btn.config(text="✓  Layout Saved!", bg="#dcfce7", fg="#166534")
            self.after(2000, lambda: self._save_btn.config(
                text="💾  Save Layout", bg="#e5e7eb", fg=FG_TEXT))
        except Exception as ex:
            self._save_btn.config(text=f"✗  Error: {ex}", bg="#fee2e2", fg="#991b1b")
            self.after(3000, lambda: self._save_btn.config(
                text="💾  Save Layout", bg="#e5e7eb", fg=FG_TEXT))

    def _load_layout(self):
        if not os.path.exists(LAYOUT_FILE):
            return {}
        try:
            with open(LAYOUT_FILE) as f:
                data = json.load(f)
            return {ip: tuple(pos) for ip, pos in data.get("positions", {}).items()}
        except Exception:
            return {}

    # ── data ──────────────────────────────────────────────────────────────────

    def _load_dialog(self):
        ft = [("CSV","*.csv")]
        if HAS_XLSX: ft.append(("Excel","*.xlsx *.xls"))
        ft.append(("All","*.*"))
        path = filedialog.askopenfilename(title="Load Device File", filetypes=ft)
        if path: self._load(path)

    def _load(self, path):
        if not path or not os.path.exists(path): return
        self._devices = load_file(path)
        self._build_map()
        self._scan()

    def _build_map(self):
        saved     = self._load_layout()
        in_memory = {ip: (d["lx"], d["ly"]) for ip, d in self._map._nodes.items()}
        preserved = {**saved, **in_memory}

        self._systems = {}
        depts      = sorted({d["Department"] for d in self._devices})
        dept_color = {d: DEPT_COLORS[i % len(DEPT_COLORS)] for i, d in enumerate(depts)}

        self._dept_cb["values"] = ["All Departments"] + depts
        self._dept_var.set("All Departments")

        positions = _tree_layout(self._source_ip, self._devices)

        node_list = [{"ip": self._source_ip, "name": "YOU", "dept": "Source",
                      "location": "This machine", "fill": C_SOURCE_CORE,
                      "lx": 0, "ly": 0, "is_source": True, "parent_ip": None,
                      "zone": ""}]
        self._systems[self._source_ip] = {"name": "YOU", "dept": "Source",
                                           "location": "", "latency": 1.0}

        for dev in self._devices:
            ip  = dev["IP"]
            pid = dev.get("Parent IP", "").strip() or None
            lx, ly = positions.get(ip, (0, 0))
            self._systems[ip] = {"name": dev["Name"], "dept": dev["Department"],
                                  "location": dev["Location"], "latency": None}
            node_list.append({
                "ip": ip, "name": dev["Name"], "dept": dev["Department"],
                "location": dev["Location"],
                "fill": dept_color[dev["Department"]],
                "lx": lx, "ly": ly, "is_source": False,
                "parent_ip": pid,
                "zone": dev.get("Zone", "").strip()
            })

        self._map.set_nodes(node_list, preserved_positions=preserved)
        self.after(150, self._map.fit_view)

    # ── ping ──────────────────────────────────────────────────────────────────

    def _schedule_scan(self):
        self._scan()
        self.after(5000, self._schedule_scan)

    def _scan(self):
        if self._scanning: return
        self._scanning = True
        threading.Thread(target=self._ping_all, daemon=True).start()

    def _ping_all(self):
        ips = [ip for ip in self._systems if ip != self._source_ip]
        futures = [self._executor.submit(lambda i: (i, do_ping(i)), ip) for ip in ips]
        for f in futures:
            try:
                ip, lat = f.result()
                self._systems[ip]["latency"] = lat
            except Exception:
                pass
        self._scanning = False
        self.after(0, self._on_done)

    def _on_done(self):
        for ip, d in self._systems.items():
            self._map.update_status(ip, d["latency"])
        self._map._redraw()
        online = sum(1 for ip, d in self._systems.items()
                     if ip != self._source_ip and d["latency"] is not None)
        total  = len(self._systems) - 1
        self._lbl_stats.config(
            text=f"Online: {online}   Offline: {total-online}   Total: {total}")
        self._lbl_time.config(
            text=f"Last scan: {datetime.now().strftime('%H:%M:%S')}")
        self._refresh_table()

    # ── table ─────────────────────────────────────────────────────────────────

    def _refresh_table(self):
        search = self._search.get().lower()
        dept_f = self._dept_var.get()
        self._tree.delete(*self._tree.get_children())
        for ip, d in self._systems.items():
            if ip == self._source_ip: continue
            if dept_f != "All Departments" and d["dept"] != dept_f: continue
            if search and not any(search in v.lower()
                                  for v in [ip, d["name"], d["dept"], d["location"]]):
                continue
            lat = d["latency"]
            self._tree.insert("","end", iid=ip,
                              values=(ip, d["name"],
                                      f"{lat} ms" if lat else "Timeout",
                                      "Online" if lat else "Offline"),
                              tags=("online" if lat else "offline",))

    def _jump_to(self, event):
        sel = self._tree.selection()
        if not sel: return
        nd = self._map._nodes.get(sel[0])
        if not nd: return
        self._map._pan  = [-nd["lx"], -nd["ly"]]
        self._map._zoom = max(self._map._zoom, 2.0)
        self._map._redraw()

    def _export(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV","*.csv")],
            initialfile="network_status.csv")
        if not path: return
        with open(path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["IP","Name","Department","Location","Latency","Status"])
            for ip, d in self._systems.items():
                if ip == self._source_ip: continue
                lat = d["latency"]
                w.writerow([ip, d["name"], d["dept"], d["location"],
                            f"{lat} ms" if lat else "Timeout",
                            "Online" if lat else "Offline"])

if __name__ == "__main__":
    app = App()
    app.mainloop()
